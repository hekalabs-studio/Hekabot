import sys
import pdfplumber
from openpyxl import Workbook

def main():
    input_path, output_path = sys.argv[1], sys.argv[2]
    wb = Workbook()
    ws = wb.active
    ws.title = "Hasil"
    row_idx = 1

    with pdfplumber.open(input_path) as pdf:
        for page in pdf.pages:
            tables = page.extract_tables()
            if tables:
                for table in tables:
                    for row in table:
                        for col_idx, cell in enumerate(row, start=1):
                            ws.cell(row=row_idx, column=col_idx, value=cell)
                        row_idx += 1
                    row_idx += 1  # baris kosong pemisah antar tabel
            else:
                text = page.extract_text() or ""
                for line in text.split("\n"):
                    if line.strip():
                        ws.cell(row=row_idx, column=1, value=line)
                        row_idx += 1

    wb.save(output_path)

if __name__ == "__main__":
    main()
